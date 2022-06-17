import pandas as pd
import openpyxl as vb
from openpyxl.styles import Font
import os
from 工具包 import OS工具

# 1.0 增
def 添加指定名称sheet在最前面(wkBookPath:str,sheetName:str):
    """

        :param wkBookPath: Workbook（工作簿）绝对路径
        :type wkBookName: str

        :param sheetName: Worksheet_Name(工作表的名称)
        :type sheetName: str

        .. note::

        添加指定名称sheet在工作簿的最前面

    """

    # 载入工作簿
    workbook = vb.load_workbook(wkBookPath)

    # 增加目标Sheet
    workbook.create_sheet(title=sheetName,index=0)

    # 保存已做处理的工作簿
    workbook.save(wkBookPath)


def 添加指定名称sheet在最后面(wkBookPath:str,sheetName:str):
    """

        :param wkBookPath: Workbook（工作簿）绝对路径
        :type wkBookPath: str

        :param sheetName: Worksheet_Name(工作表的名称)
        :type sheetName: str

        .. note::

        添加指定名称sheet在工作簿的最后面

    """

    # 载入工作簿
    workbook = vb.load_workbook(wkBookPath)

    # 增加目标Sheet
    workbook.create_sheet(title=sheetName)

    # 保存已做处理的工作簿
    workbook.save(wkBookPath)

    return None

# 2.0 删
def 删除指定名称sheet(wkBookPath,sheetName):
    """

        :param wkBookPath: Workbook（工作簿）绝对路径
        :type wkBookPath: str

        :param sheetName: Worksheet_Name(工作表的名称)
        :type sheetName: str

        .. note::

        从工作簿中删除指定名称sheet

    """

    # 载入工作簿
    workbook = vb.load_workbook(wkBookPath)

    # 删除目标Sheet
    worksheet = workbook[sheetName]
    workbook.remove(worksheet)

    # 保存已做删除处理的工作簿
    workbook.save(wkBookPath)


def 删除含指定关键字的sheet(wkBookPath,strInSheetName):
    """

        :param wkBookPath: Workbook（工作簿）绝对路径
        :type wkBookPath: str

        :param strInSheetName: 要查找的关键字
        :type strInSheetName: str

        .. note::

        从工作簿中删除含指定关键字的sheet

    """

    # 1.0 载入工作簿
    workbook = vb.load_workbook(wkBookPath)

    # 2.0 删除目标Sheet
    worksheets = workbook.worksheets

    # 3.0 循环判断名称，如果有的话，就移除
    for worksheet in worksheets:
        worksheetName = worksheet.title
        if strInSheetName in worksheetName:
            删除指定名称sheet(wkBookPath=wkBookPath,sheetName=worksheetName)

    # 4.0 保存 -> 不用保存，不然删除的Sheet又给恢复了
    # workbook.save(wkBookPath)

# 3.0 改
def 拆分一个表为多个表(wkBookPath,splitBy,toPath):
    """

        :param wkBookPath: Workbook（工作簿）绝对路径
        :type wkBookPath: str

        :param splitBy: 根据哪个字段进行拆分
        :type splitBy: str

        :param toPath: 拆分后工作簿保存的位置
        :type toPath: str

        .. note::

        将一个工作表拆分为多个工作表，并保存当当前的工作簿中

    """

    # 读取工作簿的数据
    data = pd.read_excel(wkBookPath)

    # 指定字段进行表格拆分
    depts = list(data[splitBy].drop_duplicates())

    # 开启ExcelWriter（Excel写入器）
    newbook = pd.ExcelWriter(toPath)

    for dept in depts:
        # 过滤条件
        conditions = splitBy + ' == ' + "\"" + dept + "\""

        # 拿到过滤数据
        subdata = data.query(conditions)

        # 将拿到的数据写入newbook，并指定工作表名
        subdata.to_excel(newbook,sheet_name=dept,index=None)

    # 将newbook存储到本地，并关闭开启ExcelWriter
    newbook.save()
    newbook.close()


def 拆分一个表为多个工作簿(wkBookPath,splitBy,toPath):
    """

        :param wkBookPath: Workbook（工作簿）绝对路径
        :type wkBookPath: str

        :param splitBy: 根据哪个字段进行拆分
        :type splitBy: str
        
        :param toPath: 拆分后工作簿保存的位置
        :type toPath: str

        .. note::

        将一个工作表拆分为多个工作簿，并且保存到指定的path下

    """

    # 读取工作簿的数据
    data = pd.read_excel(wkBookPath)

    # 指定字段进行表格拆分
    depts = list(data[splitBy].drop_duplicates())

    for dept in depts:
        # 过滤条件
        conditions = splitBy + ' == ' + "\"" + dept + "\""

        # 拿到过滤数据
        subdata = data.query(conditions)

        # 将拿到的数据写出到指定路径下指定的工作簿内，并指定工作表名
        subdata.to_excel(os.path.join(toPath,dept + ".xlsx"),index=None)


def 重命名工作表(wkBookPath,OWsheetName,TWsheetName):
    """

        :param wkBookPath: Workbook（工作簿）绝对路径
        :type wkBookPath: str

        :param wsheetName: 工作表原本的名称
        :type wsheetName: str


        :param TWsheetName: TargetWorksheetName，工作表修改的目标名称
        :type TWsheetName: str

        .. note::

        重命名指定名称工作簿的工作表

    """

    # 1.0 读取到工作簿
    wb = vb.load_workbook(wkBookPath)

    # 2.0 读取到指定工作表
    ws = wb[OWsheetName]

    # 3.0 修改工作表名称
    ws.title = TWsheetName

    # 4.0 保存
    wb.save(wkBookPath)


def 更改已使用单元格字体(wkBookPath,sheetName="Sheet1",fontStyle="微软雅黑",bold=False,italic=False,size=10):
    """

            :param wkBookPath: Workbook（工作簿）绝对路径
            :type wkBookPath: str

            :param sheetName: 工作表的名称，默认是“Sheet1”
            :type sheetName: str


            :param fontStyle: 字体样式，默认是“微软雅黑”
            :type fontStyle: str

            :param bold: 加粗，默认不加粗
            :type bold: str

            :param italic: 倾斜，默认不加粗
            :type italic: str

            :param size: 字体大小，默认10
            :type size: str

            .. note::

            更改所有单元格格式，根据实际情况，通常不会超过10万行，故只修改前10万行的格式

        """
    # 1.0 读取到工作簿
    wb = vb.load_workbook(wkBookPath)

    # 2.0 读取到指定工作表
    ws = wb[sheetName]

    # 3.0 字体对象
    fontObj1 = Font(name=fontStyle, bold=bold, italic=italic, size=size)

    # for eachRange in ws['A1:XFD1048576']:
    for eachRow in ws:
        for eachRange in eachRow:
            eachRange.font = fontObj1

    # 4.0 保存
    wb.save(wkBookPath)

def 保存Openpyxl的wb对象(wb,path):
    """

        :param wb: Openpyxl的
        :type wb: Openpyxl的Workbook对象

        :param path: wb要保存至的绝对路径
        :type path: str

        .. note::

        读取工作簿下的指定工作表

    """

    wb.save(path)

# 4.0 查
def 读取工作簿下的指定工作表对象_Openpyxl(wkBookPath):
    """

        :param wkBookPath: Workbook（工作簿）绝对路径
        :type wkBookPath: str

        :param wsheetName: 工作表的名称，默认是Sheet1
        :type wsheetName: str

        .. note::

        读取工作簿下的指定工作表

    """

    # 1.0 读取到工作簿
    wb = vb.load_workbook(wkBookPath)

    return wb


def 读取工作簿下的所有工作表名(wkBookPath):
    """

        :param wkBookPath: Workbook（工作簿）绝对路径
        :type wkBookPath: str

        .. note::

        读取工作簿下的所有工作表名，以数组的形式返回

    """
    # 1.0 读取到工作簿
    wb = vb.load_workbook(wkBookPath)

    # 2.0 读取到指定工作表
    wss = wb.sheetnames

    return wss


def 读取指定的Excel或CSV文件(wkBookPath:str,sheetName="Sheet1"):
    """

        :param wkBookPath: Workbook（工作簿）绝对路径
        :type wkBookPath: str

        :param sheetName: 读取的文件的表名
        :type sheetName: str

        .. note::

        读取指定的Excel或CSV文件

        sheetName  --->  如果不指定的话，默认是Sheet1

    """

    if wkBookPath.endswith(".xlsx"):

        return pd.read_excel(wkBookPath,sheet_name=sheetName)

    elif wkBookPath.endswith(".csv"):

        # 注意：csv文件不用指定sheetName
        return pd.read_csv(wkBookPath)

    else:

        print("传入的数据有误，请重新尝试，谢谢！<For any further question, please contact huangrenjun940201@outlook.com>")

        return None


def 合并单个路径下的所有CSV表(filesPath,encoding="utf-8",dataF = pd.DataFrame()):
    """

        :param filesPath: 所有csv文件父目录的绝对路径
        :type filesPath: str

        :param encoding: 编码格式，默认是utf-8
        :type encoding: str

        :param dataF: 初始Dataframe，可以自定义传入，否则初始化为空
        :type dataF: pd.Dataframe()

        .. note::

        合并单个路径下的所有CSV表，不考虑其他兄弟或子类文件夹中的文件

    """

    # 扫描最高级文件夹，获取所有的文件和文件夹信息
    files = os.scandir(filesPath)

    # 遍历获取到的files
    for file in files:

        filePath = file.path

        # 判断是文件并且是csv类型的
        if os.path.isfile(filePath) and filePath.endswith(".csv"):

            content = pd.read_csv(filePath,encoding=encoding)

            dataF = pd.concat([dataF,content])

    return dataF


def 合并指定路径下的所有CSV表_含子级目录(filesPath,encoding="utf-8",dataF = pd.DataFrame()):
    """

        :param filesPath: 所有csv文件父目录的绝对路径
        :type filesPath: str

        :param encoding: 编码格式，默认是utf-8
        :type encoding: str

        :param dataF: 初始Dataframe，可以自定义传入，否则初始化为空
        :type dataF: pd.Dataframe()

        .. note::

        合并单个路径下的所有CSV表，包含所有子孙文件夹中的CSV文件

    """

    # 1.0 扫描最高级文件夹，获取所有的文件和文件夹信息
    filesPathSubFolderArr = OS工具.获取指定文件夹下所有的子文件夹名(filesPath=filesPath)

    # 2.0 循环判断filesPathArr，对每个文件夹进行处理
    for filesPathSubFolder in filesPathSubFolderArr:

        data = 合并单个路径下的所有CSV表(filesPath=filesPathSubFolder,encoding=encoding)

        dataF = pd.concat([dataF, data])

    return dataF


def 合并单个路径下的所有Excel表(filesPath):
    """

        :param filesPath: 所有csv文件父目录的绝对路径
        :type filesPath: str

        .. note::

        合并单个路径下的所有CSV表，不考虑其他兄弟或子类文件夹中的文件

    """

    dataF = pd.DataFrame()

    # 扫描最高级文件夹，获取所有的文件和文件夹信息
    files = os.scandir(filesPath)

    # 遍历获取到的files
    for file in files:

        filePath = file.path

        if filePath.endswith(".xlsx") or filePath.endswith(".xls"):

            content = pd.read_excel(filePath)

            dataF = pd.concat([dataF,content])

    return dataF


def 合并指定路径下的所有Excel表_含子级目录(filesPath,dataF = pd.DataFrame()):
    """

        :param filesPath: 所有csv文件父目录的绝对路径
        :type filesPath: str

        :param dataF: 初始Dataframe，可以自定义传入，否则初始化为空
        :type dataF: pd.Dataframe()

        .. note::

        合并单个路径下的所有CSV表，包含所有子孙文件夹中的CSV文件

    """

    # 1.0 扫描最高级文件夹，获取所有的文件和文件夹信息
    filesPathSubFolderArr = OS工具.获取指定文件夹下所有的子文件夹名(filesPath=filesPath)

    # 2.0 循环判断filesPathArr，对每个文件夹进行处理
    for filesPathSubFolder in filesPathSubFolderArr:

        data = 合并单个路径下的所有Excel表(filesPath=filesPathSubFolder)

        dataF = pd.concat([dataF, data])

    return dataF


def 合并指定工作簿的所有sheet(wkBookPath,toPath):
    """

        :param wkBookPath: Workbook（工作簿）绝对路径
        :type wkBookPath: str

        :param toPath: 合并后工作簿保存的位置
        :type toPath: str

        .. note::

        合合并指定工作簿的所有sheet，将归总的数据存到新表中，并保存到指定的位置

    """


    data = pd.read_excel(wkBookPath)

    dataf = pd.DataFrame()

    keys = list(data.keys())
    for key in keys:
        # print(key) # 这里打印的其实是sheet的名称
        content = data[key]
        # print(content) # 这里打印的是这个sheet里面的数据
        dataf = pd.concat([dataf, content])
    return dataf


if __name__ == '__main__':
    更改已使用单元格字体(wkBookPath=r'C:\Users\huang\Desktop\VM快速部署模板.xlsx')





